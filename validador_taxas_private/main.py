import warnings
import pandas as pd
import shutil
import os
import sys
from openpyxl import load_workbook
from datetime import datetime
import contextlib

# IGNORAR WARNINGS DE USER
warnings.filterwarnings("ignore", category=UserWarning)

# CABEÇALHO
print("="*50)
print("VALIDAÇÃO DE TAXAS PRIVATE")
print("INICIANDO PROCESSAMENTO...")
print("="*50)
print()

# INPUT DE DATA
while True:
    data_str = input("|" + "-"*4 + "DATA DO CADASTRO DA TAXA: ")
    try:
        data_obj = datetime.strptime(data_str, "%d/%m/%Y")
        data_formatada = data_obj.strftime("%d-%m-%Y")
        break
    except ValueError:
        print("|" + "-"*4 + "DATA INVÁLIDA!")

print("|")
print("|" + "-"*4 + "CONVERTENDO ARQUIVO CSV EM XLSX")

# DEFINIR PASTA BASE
if getattr(sys, 'frozen', False):
    pasta_base = os.path.dirname(sys.executable)
else:
    pasta_base = os.path.dirname(os.path.abspath(__file__))

# CAMINHOS DOS ARQUIVOS
arquivo_csv = os.path.join(pasta_base, "RELAÇÃO SOMENTE SEGMENTOS.CSV")
arquivo_excel = os.path.join(pasta_base, "RELAÇÃO SOMENTE SEGMENTOS.xlsx")
arquivo_origem = arquivo_excel
arquivo_origem2 = os.path.join(pasta_base, "RELAÇÃO DE CONTRATOS COM PARCELAS.xls")
arquivo_destino = os.path.join(pasta_base, "CONFERÊNCIA CADASTRO DE TAXA PRIVATE.xlsx")
pasta_destino = os.path.join(pasta_base, "CONFERÊNCIA CADASTRO DE TAXA PRIVATE")
os.makedirs(pasta_destino, exist_ok=True)

arquivos_para_copiar = [
    arquivo_excel,
    arquivo_origem2,
    arquivo_destino
]

# FUNÇÃO PARA ABRIR ARQUIVO CSV COM TENTATIVA
while True:
    try:
        df = pd.read_csv(arquivo_csv, sep=';', encoding='latin1', on_bad_lines='skip')
        break
    except FileNotFoundError:
        print("|" + "-"*4 + "ERRO: ARQUIVO 'RELAÇÃO SOMENTE SEGMENTOS.CSV' NÃO FOI ENCONTRADO!")
        resposta = input("|" + "-"*4 + "DESEJA TENTAR NOVAMENTE? (sim/não): ").strip().lower()
        if resposta != "sim":
            print("|" + "-"*4 + "PROGRAMA ENCERRADO PELO USUÁRIO.")
            input("PRESSIONE ENTER PARA SAIR...")
            exit()
        print("|" + "-"*4 + "TENTANDO NOVAMENTE...")

# CONVERTENDO CSV PARA XLSX
df.to_excel(arquivo_excel, index=False, engine='openpyxl')
print("|")
print("|" + "-"*4 + "CSV CONVERTIDO PARA XLSX COM SUCESSO!")
print("|")
print("|" + "-"*4 + "FILTRANDO DADOS")

# FILTRAR DADOS
df_filtrado = df[~df["SegmentoContrato"].str.contains("NENHUM", na=False)]
df_filtrado.to_excel(arquivo_origem, index=False)
print("|")
print("|" + "-"*4 + "DADOS FILTRADOS COM SUCESSO!")
print("|")
print("|" + "-"*4 + "COPIANDO DADOS PARTE 01/03")

# ABRIR ARQUIVO XLSX EXISTENTE COM TENTATIVA
while True:
    try:
        df_origem = pd.read_excel(arquivo_origem, sheet_name=0)
        break
    except FileNotFoundError:
        print("|")
        print("|" + "-"*4 + "ERRO: ARQUIVO 'RELAÇÃO SOMENTE SEGMENTOS.XLSX' NÃO FOI ENCONTRADO!")
        print("|")
        resposta = input("|" + "-"*4 + "DESEJA TENTAR NOVAMENTE? (sim/não): ").strip().lower()
        if resposta != "sim":
            print("|")
            print("|" + "-"*4 + "PROGRAMA ENCERRADO PELO USUÁRIO.")
            print("|")
            input("PRESSIONE ENTER PARA SAIR...")
            exit()
        print("|")    
        print("|" + "-"*4 + "TENTANDO NOVAMENTE...")

# COPIAR DADOS PARA DESTINO (PARTE 02)
wb_destino = load_workbook(arquivo_destino)
aba_destino = wb_destino["36X"]

df_origem = df_origem.iloc[1:, :4]
linha_inicial = 1
for i, row in df_origem.iterrows():
    for j, valor in enumerate(row):
        aba_destino.cell(row=linha_inicial + i, column=j + 1, value=valor)
wb_destino.save(arquivo_destino)
print("|")
print("|" + "-"*4 + "DADOS COPIADOS COM SUCESSO!")
print("|")
print("|" + "-"*4 + "COPIANDO DADOS PARTE 02/03")

# ABRIR ARQUIVO XLS COM TENTATIVA
while True:
    try:
        with open(os.devnull, "w") as fnull:
            with contextlib.redirect_stdout(fnull):
                df_origem2 = pd.read_excel(arquivo_origem2, sheet_name=0, header=0, engine='xlrd')
        break
    except FileNotFoundError:
        print("|")
        print("|" + "-"*4 + "ERRO: ARQUIVO 'RELAÇÃO DE CONTRATOS COM PARCELAS' NÃO FOI ENCONTRADO!")
        print("|")
        resposta = input("|" + "-"*4 + "DESEJA TENTAR NOVAMENTE? (sim/não): ").strip().lower()
        if resposta != "sim":
            print("|")
            print("|" + "-"*4 + "PROGRAMA ENCERRADO PELO USUÁRIO.")
            print("|")
            input("PRESSIONE ENTER PARA SAIR...")
            exit()
        print("|")    
        print("|" + "-"*4 + "TENTANDO NOVAMENTE...")

# COPIAR DADOS PARA DESTINO (PARTE 03)
df_origem2 = df_origem2.iloc[1:, :12]
wb_destino = load_workbook(arquivo_destino)
aba_destino = wb_destino["RELAÇÃO DE CONTRATOS "]
linha_inicial = 1
for i, row in df_origem2.iterrows():
    for j, valor in enumerate(row):
        aba_destino.cell(row=linha_inicial + i, column=j + 1, value=valor)
wb_destino.save(arquivo_destino)
print("|")
print("|" + "-"*4 + "DADOS COPIADOS COM SUCESSO!")
print("|")
print("|" + "-"*4 + "COPIANDO DADOS PARTE 03/03")

# COPIAR TODOS OS ARQUIVOS PARA PASTA FINAL
print("|")
print("|" + "-"*4 + "COPIANDO ARQUIVOS PARA PASTA FINAL")
for arquivo in arquivos_para_copiar:
    destino = os.path.join(pasta_destino, os.path.basename(arquivo))
    if os.path.exists(arquivo):
        shutil.copy2(arquivo, destino)

print("|")
print("|" + "-"*4 + "TODOS OS ARQUIVOS FORAM COPIADOS PARA A PASTA FINAL!")
print("|")
print("|" + "-"*4 + "RENOMEANDO ARQUIVOS")


# RENOMEAR arquivos

arquivos_para_renomear = [
    "CONFERÊNCIA CADASTRO DE TAXA PRIVATE.xlsx",
    "RELAÇÃO DE CONTRATOS COM PARCELAS.xls",
    "RELAÇÃO SOMENTE SEGMENTOS.xlsx"
]

for nome_arquivo in arquivos_para_renomear:
    caminho_antigo = os.path.join(pasta_destino, nome_arquivo)
    nome, ext = os.path.splitext(nome_arquivo)
    novo_nome = f"{nome} - {data_formatada}{ext}"
    caminho_novo = os.path.join(pasta_destino, novo_nome)
    os.rename(caminho_antigo, caminho_novo)

print("|")
print(f"|----ARQUIVOS RENOMEADOS")


print("|")
print("|" + "-"*4 + "EXCLUINDO ARQUIVOS TEMPORÁRIOS")

pasta = pasta_base

arquivos_para_excluir = [
    "RELAÇÃO SOMENTE SEGMENTOS.CSV",
    "RELAÇÃO SOMENTE SEGMENTOS.xlsx",
    "RELAÇÃO DE CONTRATOS COM PARCELAS.xls",
    "CONFERÊNCIA CADASTRO DE TAXA PRIVATE.xlsx"
]


for nome_arquivo in arquivos_para_excluir:
    caminho = os.path.join(pasta, nome_arquivo)
    if os.path.exists(caminho):
        os.remove(caminho)
    else:
        print("|")
        
print("|" + "-"*4 + "ARQUIVOS EXCLUIDOS COM SUCESSO!")
# MANTER CONSOLE ABERTO
print("|")
print("|" + "-"*4 + "PROCESSAMENTO CONCLUÍDO!")
print("|")
input("PRESSIONE ENTER PARA SAIR...")
