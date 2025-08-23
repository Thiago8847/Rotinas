import os
import sys
import pdfplumber
import re
from openpyxl import load_workbook
import fitz  
from openpyxl.drawing.image import Image
from PIL import Image as PILImage, ImageChops
import warnings
import shutil
from openpyxl.worksheet.table import Table, TableStyleInfo

warnings.simplefilter("ignore", UserWarning)

print("="*80)
print(" "*4+"PLANILHA DE PAGAMENTO DETALHADO")
print(" "*4+"INICIANDO PROCESSAMENTO...")
print("="*80)
print()


conferencia = input(str("|" + " "*4+"DATA DA CONFERÊNCIA: "))
movimentação = input(str("|" + " "*4 +"DATA DA MOVIMENTAÇÃO: "))
venc_debito = input(str("|" + " "*4 +"DATA DE PAGAMENTO DÉBITO: "))
venc_credito = input(str("|" + " "*4 +"DATA DE PAGAMENTO crédito: "))
print()

Nome_Arquivo = "CONFERÊNCIA DETALHADO DE PAGAMENTO - MOVIMENTO.xlsx"
Tela_Web_Credito = "RELATÓRIO DE VENDAS ANALÍTICO - CRÉDITO.xlsx"
Tela_Web_Debito = "RELATÓRIO DE VENDAS ANALÍTICO - DÉBITO.xlsx"
bandeiras = "ANÁLISE BANDEIRAS - PAGAMENTO DETALHADO.xlsx"
SGP_Credito = "SGP CRÉDITO.xlsx"
SGP_Debito = "SGP DÉBITO.xlsx"
analise_master = "ANÁLISE MASTER"
analise_visa = "ANÁLISE VISA"
analise_elo = "ANÁLISE ELO"
analise_hyper = "ANÁLISE HYPER"
aba_telaweb_c = "TELA WEB - CRÉDITO"

pasta_final = "PLANILHA DE PAGAMENTO DETALHADO"

max_linha = 20000


print("="*50)
print(" "*4 +"INICIANDO TRATAMENTO ARQUIVOS - TELA WEB")
print("="*50)
print()

###################################################### ------ COPIANDO ARQUIVOS: TELA WEB PARTE 1

wb_destino = load_workbook(Nome_Arquivo)


print("|" + " "*4+"ABRINDO ARQUIVOS - TELA WEB: DÉBITO")

wb_origem1 = load_workbook(Tela_Web_Credito, keep_vba=False, keep_links=False)
Aba_origem_c = wb_origem1["Sheet1"]
Aba_Credito_TW = wb_destino["BASE CRÉDITO"]
Aba_Credito_TW2 = wb_destino["TELA WEB - CRÉDITO"]

Aba_Credito_TW2["AA3"] = venc_debito

col_web_inicio = 1
col_web_fim = 21
linhas_totais1 = Aba_origem_c.max_row
linhas_totais2 = Aba_origem_c.max_row

linha_destino_inicial = 2
col_destino_inicial = 1


print("|" + " "*4+"COPIANDO ARQUIVOS - TELA WEB: DÉBITO")

for i, row in enumerate(Aba_origem_c.iter_rows(min_row=3, max_row=linhas_totais1,
                                             min_col=col_web_inicio, max_col=col_web_fim),
                        start=0):
    for j, cell in enumerate(row, start=0):
        valor = cell.value
        Aba_Credito_TW.cell(
            row=linha_destino_inicial + i,
            column=col_destino_inicial + j,
            value=valor
        )
        Aba_Credito_TW2.cell(
            row=linha_destino_inicial + i,
            column=col_destino_inicial + j,
            value=valor
        )
print("|" + " "*4+"SALVANDO ARQUIVOS - TELA WEB: DÉBITO")

wb_destino.save(Nome_Arquivo)
wb_origem1.close()

###################################################### ------ COPIANDO ARQUIVOS: TELA WEB PARTE 2


print("|" + " "*4+"ABRINDO ARQUIVOS - TELA WEB: CRÉDITO")


wb_origem2 = load_workbook(Tela_Web_Debito, keep_vba=False, keep_links=False)
Aba_origem_d = wb_origem2["Sheet1"]
Aba_Debito_TW = wb_destino["BASE DÉBITO"]
Aba_Debito_TW2 = wb_destino["TELA WEB - DÉBITO"]
Aba_Debito_TW2["X2"] = venc_debito

linhas_totais2 = Aba_origem_d.max_row

print("|" + " "*4+"COPIANDO ARQUIVOS - TELA WEB: CRÉDITO")

for i, row in enumerate(Aba_origem_d.iter_rows(min_row=3, max_row=linhas_totais2,
                                             min_col=col_web_inicio, max_col=col_web_fim),
                        start=0):
    for j, cell in enumerate(row, start=0):
        valor = cell.value
        Aba_Debito_TW.cell(
            row=linha_destino_inicial + i,
            column=col_destino_inicial + j,
            value=valor
        )
        Aba_Debito_TW2.cell(
            row=linha_destino_inicial + i,
            column=col_destino_inicial + j,
            value=valor
        )

print("|" + " "*4+"SALVANDO ARQUIVOS - TELA WEB: CRÉDITO")



wb_destino.save(Nome_Arquivo)
wb_origem2.close()


###################################################### ------ COPIANDO ARQUIVOS: TELA WEB FIM...

print()
print("="*65)
print(" "*4 +"INICIANDO TRATAMENTO ARQUIVOS PARA ANÁLISE DE BANDEIRAS")
print("="*65)
print()

print("|" + " "*4+"CARREGANDO ARQUIVOS")

principal = load_workbook(Nome_Arquivo)
bandeiras_ = load_workbook(bandeiras, keep_vba=False, keep_links=False)

# Definir as abas
aba_principal = principal["TELA WEB - CRÉDITO"]  
aba_ban = bandeiras_["ANÁLISE MASTER"]
aba_ban2 = bandeiras_["ANÁLISE VISA"]
aba_ban3 = bandeiras_["ANÁLISE HYPER"]
aba_ban4 = bandeiras_["ANÁLISE ELO"]

# Configurações
coluInicio = 1
col_web_fim = 21
linhaDestinoInicio = 2
colulaDestinoInicio = 1

# Identificar índice da coluna "Bandeira"
cabecalho = [str(cell.value).strip().lower() for cell in aba_principal[1]]
try:
    idx_bandeira = cabecalho.index("bandeira") + 1
except ValueError:
    raise Exception("❌ Coluna 'Bandeira' não encontrada na linha 1.")

# Estrutura de filtragem
bandeiras_map = {
    "mastercard": (aba_ban, "MASTER"),
    "visa": (aba_ban2, "VISA"),
    "hiper": (aba_ban3, "HIPER"),
    "elo": (aba_ban4, "ELO")
}

# Criar contadores separados para cada bandeira
linhas_bandeiras = {key: linhaDestinoInicio for key in bandeiras_map.keys()}

print("|" + " "*4+"COPIANDO DADOS DADOS")

# Iterar pelas linhas da planilha principal
for row in aba_principal.iter_rows(min_row=2, max_row=aba_principal.max_row,
                                   min_col=coluInicio, max_col=col_web_fim):
    valor_bandeira = row[idx_bandeira - 1].value
    if not valor_bandeira:
        continue
    valor_bandeira_lower = str(valor_bandeira).lower()

    for chave, (aba_destino, nome_exibicao) in bandeiras_map.items():
        if chave in valor_bandeira_lower:
            for j, cell in enumerate(row, start=colulaDestinoInicio):
                aba_destino.cell(row=linhas_bandeiras[chave], column=j, value=cell.value)
            linhas_bandeiras[chave] += 1
            break  # só copia em uma aba

print("|" + " "*4+"DADOS COPIADOS COM SUCESSO")
print("|" + " "*4+"SALVANDO ARQUIVO")
# Salvar com segurança
try:
    bandeiras_.save(bandeiras)
except PermissionError:
    print("|" + " "*4+"❌ ERRO: O arquivo está aberto no Excel. Feche-o e tente novamente.")

bandeiras_.close()

print("|" + " "*4+"ARQUIVO SALVO!")


###################################################### ------ COPIANDO ARQUIVOS: TELA WEB FIM...



linha_destino_inicial2 = 3

print()
print("="*50)
print(" "*4 +"INICIANDO TRATAMENTO ARQUIVOS - SGP")
print("="*50)
print()


print("|" + " "*4+"ABRINDO ARQUIVOS - SGP: CRÉDITO")

wb_origem3 = load_workbook(SGP_Credito, keep_vba=False, keep_links=False)
Aba_origem_SGP_C = wb_origem3["Liquidacoes"]
Aba_credito_SGP_C = wb_destino["SGP - CRÉDITO D1"]

linhas_totais3 = Aba_origem_SGP_C.max_row

print("|" + " "*4+"COPIANDO ARQUIVOS - SGP: CRÉDITO")

for i, row in enumerate(Aba_origem_SGP_C.iter_rows(min_row=8, max_row=linhas_totais3,
                                             min_col=col_web_inicio, max_col=col_web_fim),
                        start=0):
    for j, cell in enumerate(row, start=0):
        valor = cell.value
        Aba_credito_SGP_C.cell(
            row=linha_destino_inicial2 + i,
            column=col_web_inicio + j,
            value=valor
        )

print("|" + " "*4+"SALVANDO ARQUIVOS - SGP: CRÉDITO")

wb_destino.save(Nome_Arquivo)
wb_origem3.close()


print("|" + " "*4+"ABRINDO ARQUIVOS - SGP: DÉBITO")

wb_origem4 = load_workbook(SGP_Debito, keep_vba=False, keep_links=False)
Aba_origem_SGP_D = wb_origem4["Liquidacoes"]
Aba_credito_SGP_D = wb_destino["SGP - DÉBITO D0"]

linhas_totais4 = Aba_origem_SGP_D.max_row

print("|" + " "*4+"COPIANDO ARQUIVOS - SGP: DÉBITO")

for i, row in enumerate(Aba_origem_SGP_D.iter_rows(min_row=8, max_row=linhas_totais4,
                                             min_col=col_web_inicio, max_col=col_web_fim),
                        start=0):
    for j, cell in enumerate(row, start=0):
        valor = cell.value
        Aba_credito_SGP_D.cell(
            row=linha_destino_inicial2 + i,
            column=col_destino_inicial + j,
            value=valor
        )

print("|" + " "*4+"SALVANDO ARQUIVOS - SGP: DÉBITO")

wb_destino.save(Nome_Arquivo)
wb_origem3.close()

print()
print("="*50)
print(" "*4 +"INICIANDO TRATAMENTO ARQUIVOS - PAYWARE")
print("="*50)
print()

BASE_DIR = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))

MAPEAMENTO = {
    "PAYWARE 4.14 D+2 CRÉDITO": [
        ("1 - CREDITO VISA - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", "A57"),
        ("3 - CREDITO MASTERCARD - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", "A36"),
        ("22 - CREDITO ELO - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", "A7"),
        ("28 - CREDITO HIPERCARD - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", "A25"),
    ],
    "PAYWARE 4.14 D-1 DÉBITO": [
        ("2 - DEBITO VISA - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", "A7"),
        ("4 - DEBITO MASTERCARD - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", "A22"),
        ("23 - DEBITO ELO - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", "A36"),
    ]
}

arquivos = [
    # Crédito
    {"pdf": "1 - CREDITO VISA - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", 
     "celula": "A57", "aba": "PAYWARE 4.14 D+2 CRÉDITO"},
    {"pdf": "3 - CREDITO MASTERCARD - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", 
     "celula": "A36", "aba": "PAYWARE 4.14 D+2 CRÉDITO"},
    {"pdf": "22 - CREDITO ELO - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", 
     "celula": "A7", "aba": "PAYWARE 4.14 D+2 CRÉDITO"},
    {"pdf": "28 - CREDITO HIPERCARD - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", 
     "celula": "A25", "aba": "PAYWARE 4.14 D+2 CRÉDITO"},

    # Débito
    {"pdf": "2 - DEBITO VISA - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", 
     "celula": "A36", "aba": "PAYWARE 4.14 D-1 DÉBITO"},
    {"pdf": "4 - DEBITO MASTERCARD - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", 
     "celula": "A22", "aba": "PAYWARE 4.14 D-1 DÉBITO"},
    {"pdf": "23 - DEBITO ELO - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf", 
     "celula": "A7", "aba": "PAYWARE 4.14 D-1 DÉBITO"},
]

Nome_Arquivo = "CONFERÊNCIA DETALHADO DE PAGAMENTO - MOVIMENTO.xlsx"

PDF_CREDITO = [
    "1 - CREDITO VISA - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf",
    "3 - CREDITO MASTERCARD - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf",
    "22 - CREDITO ELO - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf",
    "28 - CREDITO HIPERCARD - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf",
]
PDF_DEBITO = [
    "2 - DEBITO VISA - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf",
    "4 - DEBITO MASTERCARD - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf",
    "23 - DEBITO ELO - 4.14-Relatório Sintético de Movimentação de Estabelecimento.pdf",
]

credito = {
    "VISA": 0, 
    "MASTERCARD": 0, 
    "ELO": 0, 
    "HIPERCARD": 0
    }
debito = {
    "VISA": {
            "COMPRA": 0,
            "AJUSTE": 0,
            "VOUCHER": 0,
            "INTERNACIONAL": 0,
            "LIQUIDACAO": 0
            },
    "MASTERCARD": {
        "COMPRA": 0,
        "AJUSTE": 0, 
        "VOUCHER": 0, 
        "INTERNACIONAL": 0, 
        "LIQUIDACAO": 0},
    "ELO": {
        "COMPRA": 0, 
        "AJUSTE": 0, 
        "VOUCHER": 0, 
        "INTERNACIONAL": 0, 
        "LIQUIDACAO": 0
        },
}

def extrair_valor(texto, manter_sinal=False):
    numeros = re.findall(r"-?\d{1,3}(?:\.\d{3})*,\d{2}", texto)
    if numeros:
        valor = float(numeros[-1].replace(".", "").replace(",", "."))
        if not manter_sinal:
            valor = abs(valor)
        return valor
    return 0.0


print("|" + " "*4+"PROCESSANDO PDFs - PAYER: CRÉDITO")


for arquivo in PDF_CREDITO:
    caminho_pdf = os.path.join(BASE_DIR, arquivo)
    encontrado = False
    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if texto and "2500 - LIQUIDACAO REAIS" in texto:
                for linha in texto.split("\n"):
                    if "2500 - LIQUIDACAO REAIS" in linha:
                        valor = extrair_valor(linha)
                        encontrado = True
                        if "VISA" in arquivo.upper():
                            credito["VISA"] = valor
                        elif "MASTERCARD" in arquivo.upper():
                            credito["MASTERCARD"] = valor
                        elif "ELO" in arquivo.upper():
                            credito["ELO"] = valor
                        elif "HIPERCARD" in arquivo.upper():
                            credito["HIPERCARD"] = valor
                        break
    if not encontrado:
        if "VISA" in arquivo.upper():
            credito["VISA"] = 0
        elif "MASTERCARD" in arquivo.upper():
            credito["MASTERCARD"] = 0
        elif "ELO" in arquivo.upper():
            credito["ELO"] = 0
        elif "HIPERCARD" in arquivo.upper():
            credito["HIPERCARD"] = 0

print("|" + " "*4+"PROCESSANDO PDFs - PAYER: DÉBITO")

# Função para extrair valores de débito
def extrair_valores_debito(caminho_pdf):
    codigos = {
        "1403": "COMPRA",
        "9750": "AJUSTE",
        "1408": "VOUCHER",
        "1448": "INTERNACIONAL",
        "2500": "LIQUIDACAO"
    }
    valores = {v: 0.0 for v in codigos.values()}

    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if not texto:
                continue
            for linha in texto.split("\n"):
                for codigo, chave in codigos.items():
                    if linha.strip().startswith(codigo):
                        manter_sinal = chave in ["VOUCHER", "AJUSTE"]
                        valor = extrair_valor(linha, manter_sinal=manter_sinal)
                        valores[chave] += valor
    return valores

# Processar PDFs de débito
for arquivo in PDF_DEBITO:
    caminho_pdf = os.path.join(BASE_DIR, arquivo)
    valores = extrair_valores_debito(caminho_pdf)
    if "VISA" in arquivo.upper():
        debito["VISA"].update(valores)
    elif "MASTER" in arquivo.upper():
        debito["MASTERCARD"].update(valores)
    elif "ELO" in arquivo.upper():
        debito["ELO"].update(valores)
   
print("|" + " "*4+"COPIANDO DADOS - PAYER: CRÉDITO")

arquivo_excel = os.path.join(BASE_DIR, "CONFERÊNCIA DETALHADO DE PAGAMENTO - MOVIMENTO.xlsx")
if not os.path.exists(arquivo_excel):
    print(f"Arquivo Excel não encontrado: {arquivo_excel}")
    sys.exit(1)

wb = load_workbook(arquivo_excel)




# Crédito
aba_credito = "PAYWARE 4.14 D+2 CRÉDITO"
if aba_credito not in wb.sheetnames:
    sys.exit(1)
aba = wb[aba_credito]
aba['U10'] = credito["ELO"]
aba['U11'] = credito["HIPERCARD"]
aba['U12'] = credito["MASTERCARD"]
aba['U13'] = credito["VISA"]
aba['B4'] = movimentação
aba['C4'] = conferencia
aba['D4'] = venc_credito
# Débito
aba_debito = "PAYWARE 4.14 D-1 DÉBITO"
if aba_debito not in wb.sheetnames:
    sys.exit(1)
aba = wb[aba_debito]


aba['B4'] = movimentação
aba['C4'] = conferencia
aba['D4'] = venc_debito
# Elo
aba['W10'] = debito["ELO"]["COMPRA"]
aba['X10'] = debito["ELO"]["AJUSTE"]
aba['Y10'] = debito["ELO"]["VOUCHER"]
aba['Z10'] = debito["ELO"]["INTERNACIONAL"]
aba['AA10'] = debito["ELO"]["LIQUIDACAO"]

# Visa
aba['W12'] = debito["VISA"]["COMPRA"]
aba['X12'] = debito["VISA"]["AJUSTE"]
aba['Y12'] = debito["VISA"]["VOUCHER"]
aba['Z12'] = debito["VISA"]["INTERNACIONAL"]
aba['AA12'] = debito["VISA"]["LIQUIDACAO"]

# Mastercard
aba['W11'] = debito["MASTERCARD"]["COMPRA"]
aba['X11'] = debito["MASTERCARD"]["AJUSTE"]
aba['Y11'] = debito["MASTERCARD"]["VOUCHER"]
aba['Z11'] = debito["MASTERCARD"]["INTERNACIONAL"]
aba['AA11'] = debito["MASTERCARD"]["LIQUIDACAO"]




def trim_whitespace(img):
    bg = PILImage.new(img.mode, img.size, img.getpixel((0, 0)))
    diff = ImageChops.difference(img, bg)
    bbox = diff.getbbox()
    if bbox:
        return img.crop(bbox)
    return img

print("|" + " "*4+"COPIANDO DADOS - PAYER: DÉBITO")

# Loop processa todos os PDFs
for idx, item in enumerate(arquivos):

    doc = fitz.open(item["pdf"])
    page = doc[0]
    pix = page.get_pixmap(dpi=300)

    img_path = f"pagina_{idx}.png"
    cropped_img_path = f"pagina_cortada_{idx}.png"

    pix.save(img_path)

    img = PILImage.open(img_path)
    img = trim_whitespace(img)

    # Cortar os últimos 35% da página
    width, height = img.size
    corte_total = int(height * 0.49)
    img = img.crop((0, 0, width, corte_total))

    # Reduzir o tamanho para 60%
    new_size = (int(width * 0.35), int(corte_total * 0.35))
    img = img.resize(new_size, PILImage.Resampling.LANCZOS)

    img.save(cropped_img_path)

    # Inserir imagem na aba correta
    ws = wb[item["aba"]]
    img_excel = Image(cropped_img_path)
    ws.add_image(img_excel, item["celula"])

print()
print("="*50)
print(" "*4 +"INICIANDO PROCESSO FINAL")
print("="*50)
print()


print("|" + " "*4+"REMOVENDO ARQUIVOS TEMPORÁRIOS")

doc.close()


wb.save(arquivo_excel)

for idx in range(len(arquivos)):
    img_path = f"pagina_{idx}.png"
    cropped_img_path = f"pagina_cortada_{idx}.png"
    if os.path.exists(img_path):
        os.remove(img_path)
    if os.path.exists(cropped_img_path):
        os.remove(cropped_img_path)

for pdf in PDF_CREDITO:
    if os.path.exists(pdf):
        os.remove(pdf)

for pdf in PDF_DEBITO:
    if os.path.exists(pdf):
        os.remove(pdf)

print("|" + " "*4+"RENOMEANDO ARQUIVOS")

texto_limpo = movimentação.replace("/", "-")

todos_arquivos = [
    "CONFERÊNCIA DETALHADO DE PAGAMENTO - MOVIMENTO.xlsx",
    "RELATÓRIO DE VENDAS ANALÍTICO - CRÉDITO.xlsx",
    "RELATÓRIO DE VENDAS ANALÍTICO - DÉBITO.xlsx",
    "ANÁLISE BANDEIRAS - PAGAMENTO DETALHADO.xlsx"
]

print("|" + " "*4+"MOVENDO ARQUIVOS")

os.makedirs(pasta_final, exist_ok=True)

for arquivo in todos_arquivos:
    nome, ext = os.path.splitext(arquivo)
    novo_nome = f"{nome} - {texto_limpo}{ext}"

    origem = os.path.join(os.getcwd(), arquivo)   
    destino = os.path.join(pasta_final, novo_nome)

    shutil.move(origem, destino)

print()
print("="*50)
print(" "*4 +"PROCESSO CONCLUIDO")
print("="*50)
print()



